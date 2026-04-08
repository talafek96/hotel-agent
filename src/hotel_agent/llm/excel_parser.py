"""LLM-based Excel parser that handles any format."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import range_boundaries

from ..config import AppConfig
from ..models import Booking, Hotel, TravelerComposition
from ..utils import normalize_currency, parse_date
from .client import call_llm_json

log = logging.getLogger(__name__)


def _read_cell_value(cell: Any) -> str:
    """Read a cell's value, appending its hyperlink target if present."""
    val = cell.value
    if cell.hyperlink and cell.hyperlink.target:
        link = cell.hyperlink.target
        val = f"{val} [link: {link}]" if val else f"[link: {link}]"
    return str(val).strip() if val is not None else ""


def _read_excel_table(
    file_path: str | Path,
    sheet_name: str,
    table_name: str | None = None,
) -> tuple[list[str], list[list[Any]]]:
    """Read an Excel table and return (headers, rows) as raw strings.

    If table_name is provided, reads only the range defined by that
    named Excel table. Otherwise reads the entire sheet.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)
    ws = wb[sheet_name]

    if table_name and table_name in ws.tables:
        table = ws.tables[table_name]
        ref = table.ref
        min_col, min_row, max_col, max_row = range_boundaries(ref)

        # Also check column A for hyperlinks (often outside the table)
        include_col_a = (min_col or 1) > 1

        if_rows_data: list[list[str]] = []
        for _row_idx, row in enumerate(
            ws.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=1 if include_col_a else min_col,
                max_col=max_col,
                values_only=False,
            )
        ):
            if_rows_data.append([_read_cell_value(cell) for cell in row])

        # First row = headers
        headers = if_rows_data[0] if if_rows_data else []
        data_rows = if_rows_data[1:] if len(if_rows_data) > 1 else []

        # Filter out completely empty rows and the totals row
        data_rows = [r for r in data_rows if any(v for v in r)]

        wb.close()
        return headers, data_rows
    else:
        # Read the entire sheet
        rows_data: list[list[str]] = []
        for raw_row in ws.iter_rows(values_only=False):
            rows_data.append([_read_cell_value(cell) for cell in raw_row])

        # Try to find header row (first row with multiple non-empty cells)
        header_idx = 0
        for i, parsed_row in enumerate(rows_data):
            non_empty = sum(bool(v) for v in parsed_row)
            if non_empty >= 3:
                header_idx = i
                break

        headers = rows_data[header_idx]
        data_rows = [r for r in rows_data[header_idx + 1 :] if any(v for v in r)]
        wb.close()
        return headers, data_rows


def _format_table_for_llm(headers: list[str], rows: list[list[Any]]) -> str:
    """Format an Excel table as a readable text table for the LLM."""
    lines = []

    # Header line with column indices
    header_line = " | ".join(f"[{i}] {h}" for i, h in enumerate(headers))
    lines.append(f"HEADERS: {header_line}")
    lines.append("-" * 80)

    for row_idx, row in enumerate(rows):
        parts = []
        for col_idx, val in enumerate(row):
            if val:
                col_name = headers[col_idx] if col_idx < len(headers) else f"col{col_idx}"
                parts.append(f"{col_name}: {val}")
        if parts:
            lines.append(f"ROW {row_idx + 1}: {' | '.join(parts)}")

    return "\n".join(lines)


PARSE_SYSTEM_PROMPT = """You are a data extraction assistant. You parse hotel booking data from Excel spreadsheets.

IMPORTANT RULES:
- Extract ALL hotel entries from the data, do not skip any rows
- Dates must be in YYYY-MM-DD format
- Prices should be numeric (no currency symbols)
- Currency MUST be a 3-letter ISO 4217 code: JPY, USD, EUR, ILS, GBP, INR, etc.
  Do NOT use symbols (¥, $, ₪, €) or words (yen, euro, shekel) — always use the 3-letter code.
  Infer the currency from column headers, symbols, or country context.
- CURRENCY SANITY CHECKS — apply these before finalizing each entry:
  1. If a price appears identically in columns for different currencies, the user likely pasted it into the wrong column. Use the hotel's country and price magnitude to determine the true currency.
  2. Cross-check price vs currency: a hotel night typically costs $20-$1000 equivalent. If the price-per-night in the assigned currency converts to less than ~$10 or more than ~$5000 USD, the currency is probably wrong. Re-examine the data.
  3. Cross-check currency vs country: a hotel in India is unlikely priced in JPY; a hotel in Japan is unlikely priced in ILS. If the currency doesn't match the country, double-check. It may be correct (booked through a foreign OTA), but flag anything suspicious in the "notes" field.
  4. If you cannot confidently determine the currency, set currency to null and explain the ambiguity in "notes" so the user can correct it.
- If a booking reference or order number is present, extract it
- If cancellation info is present, extract the deadline date
- Platform names should be normalized: "Agoda", "Booking.com", "Hotels.com", "Expedia", "Direct"
- Extract hyperlinks/URLs if present — they may be booking confirmation links
- If a row seems like notes/comments rather than a hotel entry, skip it
- The data might be in Hebrew, Japanese, or English — handle all languages
- ALL fields below are OPTIONAL. Fill in what you can find in the data, omit or set to null anything not available. Do NOT fail or complain about missing columns.

Return a JSON object with this structure (include only fields you can populate):
{
  "hotels": [
    {
      "name": "Hotel Name (REQUIRED — skip row if no hotel name)",
      "city": "City",
      "country": "Country (infer from context if not explicit)",
      "address": "Full address if available",
      "url": "Hotel page URL if found",
      "booking_url": "Booking confirmation link if found",
      "check_in": "YYYY-MM-DD",
      "check_out": "YYYY-MM-DD",
      "room_type": "e.g. Deluxe Twin, Standard Double",
      "price": 12345,
      "currency": "JPY (MUST be 3-letter ISO code, not a symbol)",
      "platform": "Agoda",
      "booking_reference": "reference number if found",
      "is_cancellable": true,
      "cancellation_deadline": "YYYY-MM-DD or null",
      "breakfast_included": false,
      "dinner_included": false,
      "extras": "any extras noted",
      "notes": "any relevant notes"
    }
  ],
  "inferred_context": {
    "trip_destination": "country/region",
    "trip_dates": "overall date range",
    "currency_primary": "primary currency used",
    "notes": "any other observations about the data"
  }
}"""


def parse_excel_with_llm(
    config: AppConfig,
    file_path: str | Path,
    sheet_name: str,
    table_name: str | None = None,
    existing_bookings: list[dict] | None = None,
) -> list[dict]:
    """Parse an Excel file using LLM to extract hotel bookings.

    Parameters
    ----------
    existing_bookings:
        Optional list of ``{"booking_id": int, "hotel": str, "city": str,
        "check_in": str, "check_out": str, "booking_reference": str,
        "platform": str}`` dicts from the current database.  Passed to
        the LLM so it can match rows against existing records.

    Returns a list of parsed hotel booking dicts.
    """
    log.info(f"Reading Excel: {file_path}, sheet={sheet_name}, table={table_name}")

    headers, rows = _read_excel_table(file_path, sheet_name, table_name)
    table_text = _format_table_for_llm(headers, rows)

    log.info(f"Table has {len(headers)} columns, {len(rows)} data rows")
    log.info(f"Table text size: {len(table_text)} chars")

    existing_section = ""
    if existing_bookings:
        lines = []
        for eb in existing_bookings:
            parts = [f"id={eb['booking_id']}"]
            for k in ("hotel", "city", "check_in", "check_out", "booking_reference", "platform"):
                v = eb.get(k)
                if v:
                    parts.append(f"{k}={v}")
            lines.append("  " + ", ".join(parts))
        existing_section = (
            "\n\nEXISTING BOOKINGS IN DATABASE (match against these when possible):\n"
            + "\n".join(lines)
            + "\n\nFor each hotel entry, if it matches an existing booking above, "
            'include "existing_booking_id": <id> in the output. '
            "Match by booking reference first, then by hotel name + dates."
        )

    prompt = f"""Parse the following hotel booking data from an Excel spreadsheet.

EXCEL DATA:
{table_text}{existing_section}

Extract every hotel booking entry and return as JSON following the schema in your instructions."""

    result = call_llm_json(
        config=config,
        prompt=prompt,
        system_prompt=PARSE_SYSTEM_PROMPT,
    )

    hotels: list[dict[Any, Any]] = result.get("hotels", [])
    context = result.get("inferred_context", {})

    log.info(f"LLM extracted {len(hotels)} hotels")
    if context:
        log.info(f"Inferred context: {context}")

    return hotels


def excel_to_models(
    parsed: list[dict],
    default_travelers: TravelerComposition | None = None,
) -> list[tuple[Hotel, Booking]]:
    """Convert LLM-parsed dicts to Hotel + Booking model pairs."""
    travelers = default_travelers or TravelerComposition()
    results = []

    for entry in parsed:
        hotel = Hotel(
            name=entry.get("name", ""),
            city=entry.get("city", ""),
            country=entry.get("country", ""),
            address=entry.get("address", ""),
            url=entry.get("url", ""),
            platform=entry.get("platform", ""),
        )

        price = float(entry.get("price") or 0)
        currency = normalize_currency(entry.get("currency"))
        notes_parts = [entry.get("notes", "")]

        # Format-agnostic sanity warnings (based on parsed output only)
        warnings = _check_price_sanity(hotel.name, hotel.country, price, currency, entry)
        if warnings:
            notes_parts.extend(warnings)
            for w in warnings:
                log.warning("Import warning for %s: %s", hotel.name, w)

        booking = Booking(
            check_in=parse_date(entry.get("check_in")),
            check_out=parse_date(entry.get("check_out")),
            travelers=travelers,
            room_type=entry.get("room_type", ""),
            booked_price=price,
            currency=currency,
            is_cancellable=bool(entry.get("is_cancellable", False)),
            cancellation_deadline=parse_date(entry.get("cancellation_deadline")),
            breakfast_included=bool(entry.get("breakfast_included", False)),
            dinner_included=bool(entry.get("dinner_included", False)),
            platform=entry.get("platform", ""),
            booking_reference=entry.get("booking_reference", ""),
            booking_url=entry.get("booking_url", "") or entry.get("url", ""),
            extras=entry.get("extras", ""),
            notes=" | ".join(p for p in notes_parts if p),
        )

        results.append((hotel, booking))

    return results


# Approximate USD equivalents for a rough magnitude check.
_ROUGH_USD_RATE: dict[str, float] = {
    "USD": 1.0,
    "EUR": 1.08,
    "GBP": 1.27,
    "ILS": 0.28,
    "JPY": 0.0067,
    "INR": 0.012,
    "KRW": 0.00073,
    "THB": 0.029,
    "CNY": 0.14,
    "TWD": 0.031,
    "MXN": 0.058,
    "BRL": 0.19,
    "AUD": 0.65,
    "CAD": 0.74,
    "CHF": 1.12,
    "SGD": 0.74,
    "HKD": 0.13,
    "NZD": 0.60,
    "SEK": 0.097,
    "NOK": 0.093,
    "DKK": 0.14,
    "PLN": 0.25,
    "CZK": 0.043,
    "HUF": 0.0027,
    "TRY": 0.031,
    "ZAR": 0.055,
    "MYR": 0.22,
    "PHP": 0.018,
    "VND": 0.000040,
    "LKR": 0.0033,
}

# Country name → expected local currencies (lowercase).
_COUNTRY_CURRENCIES: dict[str, set[str]] = {
    "japan": {"JPY"},
    "india": {"INR", "USD"},
    "israel": {"ILS", "USD"},
    "sri lanka": {"LKR", "USD"},
    "austria": {"EUR"},
    "germany": {"EUR"},
    "france": {"EUR"},
    "italy": {"EUR"},
    "spain": {"EUR"},
    "united kingdom": {"GBP"},
    "thailand": {"THB", "USD"},
    "south korea": {"KRW"},
    "china": {"CNY"},
    "taiwan": {"TWD"},
    "united states": {"USD"},
    "australia": {"AUD"},
    "canada": {"CAD"},
    "mexico": {"MXN"},
    "brazil": {"BRL"},
    "turkey": {"TRY", "EUR", "USD"},
    "switzerland": {"CHF", "EUR"},
}


def _check_price_sanity(
    hotel_name: str,
    country: str,
    price: float,
    currency: str,
    entry: dict,
) -> list[str]:
    """Return warning strings for suspicious price/currency combos.

    Checks are purely on the LLM's parsed output — no assumptions about
    the original table format.
    """
    warnings: list[str] = []
    if price <= 0:
        return warnings

    nights = 1
    check_in = parse_date(entry.get("check_in"))
    check_out = parse_date(entry.get("check_out"))
    if check_in and check_out and check_out > check_in:
        nights = (check_out - check_in).days

    ppn = price / nights
    rate = _ROUGH_USD_RATE.get(currency)

    # 1. Price magnitude check
    if rate:
        usd_per_night = ppn * rate
        if usd_per_night < 12:
            warnings.append(
                f"SUSPICIOUS PRICE: {price:,.0f} {currency} "
                f"(~${usd_per_night:.0f}/night) seems too low — verify currency"
            )
        elif usd_per_night > 8000:
            warnings.append(
                f"SUSPICIOUS PRICE: {price:,.0f} {currency} "
                f"(~${usd_per_night:,.0f}/night) seems too high — verify currency"
            )

    # 2. Country/currency mismatch check
    if country:
        expected = _COUNTRY_CURRENCIES.get(country.strip().lower())
        if expected and currency not in expected:
            warnings.append(
                f"CURRENCY/COUNTRY MISMATCH: {currency} for a hotel in {country} "
                f"(expected {'/'.join(sorted(expected))})"
            )

    return warnings
